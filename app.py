from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
import os
import tempfile
import traceback
import hashlib
import datetime
import pandas as pd
import logging
import json
import uuid
import random
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Excel storage paths
EXCEL_DATA_DIR = 'excel_data'
USERS_EXCEL = os.path.join(EXCEL_DATA_DIR, 'users.xlsx')
API_KEYS_EXCEL = os.path.join(EXCEL_DATA_DIR, 'api_keys.xlsx')
PROCESSING_SESSIONS_EXCEL = os.path.join(EXCEL_DATA_DIR, 'processing_sessions.xlsx')
PROCESSED_DATA_EXCEL = os.path.join(EXCEL_DATA_DIR, 'processed_data.xlsx')
REDACTIONS_LOG_EXCEL = os.path.join(EXCEL_DATA_DIR, 'redactions_log.xlsx') 
# Import the PII/PHI detection modules
try:
    from extractor import UniversalTextExtractor
    from pii_detector import CleanPIIDetector
    from replacer import SimplePIIFaker
    PII_DETECTION_AVAILABLE = True
    print("✅ PII/PHI detection modules loaded successfully")
except ImportError as e:
    print(f"⚠️ PII/PHI detection modules not available: {e}")
    PII_DETECTION_AVAILABLE = False

# Initialize AI components if available
if PII_DETECTION_AVAILABLE:
    try:
        extractor = UniversalTextExtractor()
        detector = CleanPIIDetector()
        faker = SimplePIIFaker()
        logger.info("AI components initialized successfully")
    except Exception as e:
        logger.error(f"Failed to initialize AI components: {e}")
        PII_DETECTION_AVAILABLE = False

# EXCEL STORAGE SYSTEM INITIALIZATION
def init_excel_storage():
    """Initialize Excel files for data storage"""
    os.makedirs(EXCEL_DATA_DIR, exist_ok=True)
    
    # Initialize Users Excel
    if not os.path.exists(USERS_EXCEL):
        users_df = pd.DataFrame(columns=[
            'id', 'name', 'email', 'password_hash', 'role', 'organization', 
            'department', 'created_at', 'last_login', 'is_active'
        ])
        users_df.to_excel(USERS_EXCEL, index=False)
        logger.info("Created users.xlsx")
    
    # Initialize API Keys Excel
    if not os.path.exists(API_KEYS_EXCEL):
        api_keys_df = pd.DataFrame(columns=[
            'id', 'user_id', 'api_key', 'name', 'permissions', 
            'active', 'created_at', 'last_used', 'usage_count'
        ])
        api_keys_df.to_excel(API_KEYS_EXCEL, index=False)
        logger.info("Created api_keys.xlsx")
    
    # Initialize Processing Sessions Excel
    if not os.path.exists(PROCESSING_SESSIONS_EXCEL):
        sessions_df = pd.DataFrame(columns=[
            'id', 'user_id', 'session_name', 'files_processed', 'pii_items', 
            'phi_items', 'processing_time', 'created_at', 'status', 'notes'
        ])
        sessions_df.to_excel(PROCESSING_SESSIONS_EXCEL, index=False)
        logger.info("Created processing_sessions.xlsx")
    
    # Initialize Processed Data Excel
    if not os.path.exists(PROCESSED_DATA_EXCEL):
        processed_df = pd.DataFrame(columns=[
            'id', 'session_id', 'user_id', 'processed_at', 'original_filename',
            'file_size', 'file_type', 'pii_count', 'phi_count', 'processing_status'
        ])
        processed_df.to_excel(PROCESSED_DATA_EXCEL, index=False)
        logger.info("Created processed_data.xlsx")
    
    # Initialize Redactions Log Excel - NEW
    if not os.path.exists(REDACTIONS_LOG_EXCEL):
        redactions_df = pd.DataFrame(columns=[
            'id', 'session_id', 'user_id', 'filename', 'processed_at',
            # PII columns (without PII_ prefix)
            'Hospital_ID', 'Patient_Name', 'Policy_Number', 'date_of_birth', 
            'email', 'health_insurance', 'phone_number',
            # PHI columns (without PHI_ prefix)
            'Age', 'allergy', 'blood_group', 'drug', 
            'gender', 'medical_condition', 'surgery', 'symptom'
        ])
        redactions_df.to_excel(REDACTIONS_LOG_EXCEL, index=False)
        logger.info("Created redactions_log.xlsx")
        
    logger.info("Excel storage initialization completed successfully")

# Excel Helper Functions
def read_excel_data(file_path):
    """Read data from Excel file"""
    try:
        if os.path.exists(file_path):
            return pd.read_excel(file_path)
        return pd.DataFrame()
    except Exception as e:
        logger.error(f"Error reading {file_path}: {e}")
        return pd.DataFrame()

def write_excel_data(df, file_path):
    """Write data to Excel file"""
    try:
        df.to_excel(file_path, index=False)
        return True
    except Exception as e:
        logger.error(f"Error writing {file_path}: {e}")
        return False

def generate_id():
    """Generate a unique ID"""
    return str(uuid.uuid4())

def hash_password(password):
    """Hash a password"""
    return generate_password_hash(password, method='pbkdf2:sha256')

def check_password_hash_func(hash_val, password):
    """Check if password matches hash"""
    return check_password_hash(hash_val, password)

# Initialize Excel storage on startup
init_excel_storage()

@app.route('/')
def index():
    return render_template('index.html')

# USER REGISTRATION WITH EXCEL STORAGE
@app.route('/register', methods=['POST'])
def register():
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        department = data.get('department', '').strip()
        organization = data.get('organization', 'Hospital').strip()
        
        # Validation
        if not name or not email or not password:
            return jsonify({'success': False, 'detail': 'All fields are required'}), 400
        
        if len(name) < 2:
            return jsonify({'success': False, 'detail': 'Name must be at least 2 characters'}), 400
        
        if len(password) < 6:
            return jsonify({'success': False, 'detail': 'Password must be at least 6 characters'}), 400
        
        # Read existing users
        users_df = read_excel_data(USERS_EXCEL)
        
        # Check if user exists
        if not users_df.empty and email in users_df['email'].values:
            return jsonify({'success': False, 'detail': 'Email already registered'}), 400
        
        # Create new user
        user_id = generate_id()
        password_hash = hash_password(password)
        created_at = datetime.datetime.utcnow().isoformat()
        
        new_user = {
            'id': user_id,
            'name': name,
            'email': email,
            'password_hash': password_hash,
            'role': 'user',
            'organization': organization,
            'department': department,
            'created_at': created_at,
            'last_login': '',
            'is_active': 1
        }
        
        # Add to DataFrame and save
        users_df = pd.concat([users_df, pd.DataFrame([new_user])], ignore_index=True)
        if not write_excel_data(users_df, USERS_EXCEL):
            return jsonify({'success': False, 'detail': 'Failed to save user data'}), 500
        
        # Generate default API key
        api_keys_df = read_excel_data(API_KEYS_EXCEL)
        api_key_id = generate_id()
        api_key = f"ak_{generate_id().replace('-', '')[:32]}"
        
        new_api_key = {
            'id': api_key_id,
            'user_id': user_id,
            'api_key': api_key,
            'name': 'Default API Key',
            'permissions': 'read,write',
            'active': 1,
            'created_at': created_at,
            'last_used': '',
            'usage_count': 0
        }
        
        api_keys_df = pd.concat([api_keys_df, pd.DataFrame([new_api_key])], ignore_index=True)
        write_excel_data(api_keys_df, API_KEYS_EXCEL)
        
        logger.info(f"User registered: {email} from {organization}")
        
        return jsonify({
            'success': True,
            'message': 'Account created successfully!',
            'user': {
                'id': user_id,
                'name': name,
                'email': email,
                'organization': organization,
                'department': department,
                'createdAt': created_at
            }
        })
        
    except Exception as e:
        logger.error(f"Registration error: {str(e)}")
        return jsonify({'success': False, 'detail': 'Registration failed'}), 500

# USER LOGIN WITH EXCEL STORAGE
@app.route('/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        
        if not email or not password:
            return jsonify({'success': False, 'detail': 'Email and password required'}), 400
        
        # Read users data
        users_df = read_excel_data(USERS_EXCEL)
        
        if users_df.empty:
            return jsonify({'success': False, 'detail': 'Invalid email or password'}), 401
        
        # Find user
        user_row = users_df[users_df['email'] == email]
        if user_row.empty or not check_password_hash_func(user_row.iloc[0]['password_hash'], password):
            return jsonify({'success': False, 'detail': 'Invalid email or password'}), 401
        
        user_data = user_row.iloc[0]
        
        if not user_data['is_active']:
            return jsonify({'success': False, 'detail': 'Account is deactivated'}), 401
        
        # Update last login
        last_login = datetime.datetime.utcnow().isoformat()
        users_df.loc[users_df['email'] == email, 'last_login'] = last_login
        write_excel_data(users_df, USERS_EXCEL)
        
        # Get user's API keys
        api_keys_df = read_excel_data(API_KEYS_EXCEL)
        user_keys = api_keys_df[
            (api_keys_df['user_id'] == user_data['id']) & 
            (api_keys_df['active'] == 1)
        ]
        
        api_keys = []
        for _, key in user_keys.iterrows():
            api_keys.append({
                'id': key['id'],
                'apiKey': key['api_key'],
                'name': key['name'],
                'createdAt': key['created_at'],
                'usageCount': key['usage_count'] if pd.notna(key['usage_count']) else 0
            })
        
        logger.info(f"User logged in: {email}")
        
        return jsonify({
            'success': True,
            'message': 'Login successful!',
            'user': {
                'id': user_data['id'],
                'name': user_data['name'],
                'email': user_data['email'],
                'organization': user_data['organization'] if pd.notna(user_data['organization']) else 'Hospital',
                'department': user_data['department'] if pd.notna(user_data['department']) else '',
                'createdAt': user_data['created_at']
            },
            'apiKeys': api_keys
        })
        
    except Exception as e:
        logger.error(f"Login error: {str(e)}")
        return jsonify({'success': False, 'detail': 'Login failed'}), 500

# ENHANCED FILE PROCESSING WITH PII/PHI EXCEL OUTPUT AND REDACTIONS LOG
@app.route('/api/process-files', methods=['POST'])
def process_files():
    try:
        logger.info("Processing files with PII/PHI detection and Excel output")
        
        user_id = request.form.get('userId', 'anonymous')
        files = request.files.getlist('files')
        
        if not files:
            return jsonify({'success': False, 'detail': 'No files uploaded'}), 400
        
        session_id = generate_id()
        session_name = f"Session_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        created_at = datetime.datetime.utcnow().isoformat()
        
        processed_files = []
        pii_phi_rows = []  # For immediate download
        redactions_log_rows = []  # For persistent audit log
        total_pii_items = 0
        total_phi_items = 0
        
        for file in files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                logger.info(f"Processing file: {filename}")
                
                fd, temp_path = tempfile.mkstemp(suffix=os.path.splitext(filename)[1])
                os.close(fd)
                
                try:
                    file.save(temp_path)
                    file_size = os.path.getsize(temp_path)
                    
                    # REAL PII/PHI PROCESSING
                    if PII_DETECTION_AVAILABLE:
                        # Extract PII and PHI
                        pii_results = detector.extract_pii_from_file(temp_path)
                        phi_results = detector.extract_phi_from_file(temp_path)
                        
                        # Generate fake PII
                        if not pii_results.get("error"):
                            fake_pii = faker.replace_pii_json(pii_results)
                        else:
                            fake_pii = {}
                        
                        # Keep real PHI
                        if not phi_results.get("error"):
                            real_phi = phi_results
                        else:
                            real_phi = {}
                    else:
                        # Fallback sample data
                        fake_pii = {
                            "Patient Name": f"Patient_{random.randint(1000, 9999)}",
                            "phone number": f"+1-555-{random.randint(1000, 9999)}",
                            "email": f"patient{random.randint(100, 999)}@example.com",
                            "Address": f"{random.randint(100, 999)} Main St, City, State {random.randint(10000, 99999)}",
                            "Hospital ID": f"HOSP-{random.randint(1000000, 9999999)}",
                            "Policy Number": f"POL-{random.randint(10000000, 99999999)}",
                            "date of birth": f"{random.randint(1, 28)} {random.choice(['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun'])} {random.randint(1970, 2000)}",
                            "health insurance": random.choice(["Health Insurance Corp", "Medical Care Plus", "Wellness Insurance"])
                        }
                        real_phi = {
                            "medical condition": random.choice(["Hypertension", "Diabetes", "Asthma", "Arthritis"]),
                            "Age": str(random.randint(25, 80)),
                            "gender": random.choice(["Male", "Female"]),
                            "medication": random.choice(["Lisinopril", "Metformin", "Aspirin", "Ibuprofen"]),
                            "allergy": random.choice(["Penicillin", "Peanuts", "No known allergies"]),
                            "blood group": random.choice(["A+", "B+", "O+", "AB+", "A-", "B-", "O-", "AB-"]),
                            "surgery": random.choice(["Appendectomy", "No surgeries", "Knee replacement"]),
                            "symptom": random.choice(["Headache", "Fatigue", "No symptoms", "Joint pain"])
                        }
                    
                    pii_count = len(fake_pii) if fake_pii else 0
                    phi_count = len(real_phi) if real_phi else 0
                    total_pii_items += pii_count
                    total_phi_items += phi_count
                    
                    # Create individual file data (without prefixes) - ONE ROW PER FILE
                    file_data = {
                        'filename': filename,
                        'Hospital_ID': fake_pii.get('Hospital ID', f'HOSP-{random.randint(1000000, 9999999)}'),
                        'Patient_Name': fake_pii.get('Patient Name', 'John Doe'),
                        'Policy_Number': fake_pii.get('Policy Number', f'POL-{random.randint(10000000, 99999999)}'),
                        'date_of_birth': fake_pii.get('date of birth', '15 March 1990'),
                        'email': fake_pii.get('email', 'patient@example.com'),
                        'health_insurance': fake_pii.get('health insurance', 'Health Insurance Corp'),
                        'phone_number': fake_pii.get('phone number', '+91-9876543210'),
                        'Age': real_phi.get('Age', '35'),
                        'allergy': real_phi.get('allergy', 'No known allergies'),
                        'blood_group': real_phi.get('blood group', 'O+'),
                        'drug': real_phi.get('medication', 'No medications'),
                        'gender': real_phi.get('gender', 'Male'),
                        'medical_condition': real_phi.get('medical condition', 'No conditions'),
                        'surgery': real_phi.get('surgery', 'No surgeries'),
                        'symptom': real_phi.get('symptom', 'No symptoms')
                    }
                    
                    # Add to session results
                    pii_phi_rows.append(file_data)
                    
                    # Add to redactions log (persistent audit trail)
                    redaction_record = {
                        'id': generate_id(),
                        'session_id': session_id,
                        'user_id': user_id,
                        'filename': filename,
                        'processed_at': created_at,
                        # Spread all file data except filename (which is already included)
                        **{k: v for k, v in file_data.items() if k != 'filename'}
                    }
                    redactions_log_rows.append(redaction_record)
                    
                    processed_files.append({
                        'id': generate_id(),
                        'filename': filename,
                        'size': file_size,
                        'pii_items': pii_count,
                        'phi_items': phi_count,
                        'status': 'completed'
                    })
                    
                except Exception as e:
                    logger.error(f"Error processing {filename}: {e}")
                    processed_files.append({
                        'id': generate_id(),
                        'filename': filename,
                        'error': str(e),
                        'status': 'failed'
                    })
                
                finally:
                    try:
                        os.remove(temp_path)
                    except:
                        pass
        
        # After processing all files, append to redactions log
        if redactions_log_rows:
            redactions_df = read_excel_data(REDACTIONS_LOG_EXCEL)
            new_redactions_df = pd.DataFrame(redactions_log_rows)
            redactions_df = pd.concat([redactions_df, new_redactions_df], ignore_index=True)
            write_excel_data(redactions_df, REDACTIONS_LOG_EXCEL)
            logger.info(f"Added {len(redactions_log_rows)} records to redactions log")
        
        processing_time = len(processed_files) * 0.8 + (total_pii_items + total_phi_items) * 0.02
        
        # Save processing session to Excel
        sessions_df = read_excel_data(PROCESSING_SESSIONS_EXCEL)
        new_session = {
            'id': session_id,
            'user_id': user_id,
            'session_name': session_name,
            'files_processed': len(processed_files),
            'pii_items': total_pii_items,
            'phi_items': total_phi_items,
            'processing_time': round(processing_time, 2),
            'created_at': created_at,
            'status': 'completed',
            'notes': f"Processed {len(processed_files)} files with redactions log"
        }
        
        sessions_df = pd.concat([sessions_df, pd.DataFrame([new_session])], ignore_index=True)
        write_excel_data(sessions_df, PROCESSING_SESSIONS_EXCEL)
        
        # Save processed files data
        processed_df = read_excel_data(PROCESSED_DATA_EXCEL)
        for file_info in processed_files:
            if file_info.get('status') == 'completed':
                processed_record = {
                    'id': generate_id(),
                    'session_id': session_id,
                    'user_id': user_id,
                    'processed_at': created_at,
                    'original_filename': file_info['filename'],
                    'file_size': file_info['size'],
                    'file_type': os.path.splitext(file_info['filename'])[1],
                    'pii_count': file_info['pii_items'],
                    'phi_count': file_info['phi_items'],
                    'processing_status': 'completed'
                }
                processed_df = pd.concat([processed_df, pd.DataFrame([processed_record])], ignore_index=True)
        
        write_excel_data(processed_df, PROCESSED_DATA_EXCEL)
        
        logger.info(f"Processing completed: {session_id} - {len(pii_phi_rows)} files processed")
        
        return jsonify({
            'success': True,
            'results': {
                'sessionId': session_id,
                'filesProcessed': len(processed_files),
                'piiItemsRemoved': total_pii_items,
                'phiItemsDetected': total_phi_items,
                'processingTime': round(processing_time, 2),
                'processedFiles': processed_files,
                'piiPhiData': pii_phi_rows  # Each file as separate row
            }
        })
        
    except Exception as e:
        logger.error(f"Processing error: {str(e)}")
        return jsonify({'success': False, 'detail': f'Processing failed: {str(e)}'}), 500
        
# EXCEL DOWNLOAD WITH PROPER PII/PHI STRUCTURE
@app.route('/api/download-results', methods=['POST'])
def download_results():
    try:
        data = request.get_json()
        pii_phi_data = data.get('piiPhiData', [])
        
        if not pii_phi_data:
            return jsonify({'success': False, 'detail': 'No data to download'}), 400
        
        # Create DataFrame without column prefixes
        df = pd.DataFrame(pii_phi_data)
        
        # Ensure columns are in the right order (filename first, then data)
        column_order = ['filename', 'Hospital_ID', 'Patient_Name', 'Policy_Number', 
                       'date_of_birth', 'email', 'health_insurance', 'phone_number',
                       'Age', 'allergy', 'blood_group', 'drug', 'gender', 
                       'medical_condition', 'surgery', 'symptom']
        
        df = df.reindex(columns=column_order, fill_value='')
        
        # Create Excel file with formatting
        fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        
        try:
            with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
                # Write main data
                df.to_excel(writer, sheet_name='PII_PHI_Data', index=False)
                
                # Format the worksheet
                worksheet = writer.sheets['PII_PHI_Data']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 40)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add header formatting
                from openpyxl.styles import Font, PatternFill
                header_font = Font(bold=True, size=12)
                basic_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')  # Light green
                pii_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')    # Light blue
                phi_fill = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')    # Light orange
                
                # Format headers
                # Remove the undefined basic_columns reference:
                for cell in worksheet[1]:  # Header row
                    cell.font = header_font
                    if cell.value and cell.value.startswith('PII_'):
                        cell.fill = pii_fill
                    elif cell.value and cell.value.startswith('PHI_'):
                        cell.fill = phi_fill
                
                # Add a summary sheet
                # Count PII and PHI columns dynamically:
            pii_cols = [col for col in df.columns if col.startswith('PII_')]
            phi_cols = [col for col in df.columns if col.startswith('PHI_')]

            summary_data = {
                'Metric': [
                    'Total Files Processed',
                    'Total PII Types Found',
                    'Total PHI Types Found',
                    'Processing Date'
                ],
                'Value': [
                    len(df),
                    len(pii_cols),
                    len(phi_cols),
                    datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            }
                
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format summary sheet
            summary_worksheet = writer.sheets['Summary']
            for column in summary_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                summary_worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format summary headers
                for cell in summary_worksheet[1]:
                    cell.font = header_font
                    cell.fill = basic_fill
            
            download_name = f'PII_PHI_Results_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            return send_file(
                temp_path,
                as_attachment=True,
                download_name=download_name,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            try:
                os.remove(temp_path)
            except:
                pass
            raise e
            
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return jsonify({'success': False, 'detail': 'Download failed'}), 500

# API KEY GENERATION WITH EXCEL STORAGE
@app.route('/api/generate-key', methods=['POST'])
def generate_api_key():
    try:
        data = request.get_json()
        user_id = data.get('userId')
        key_name = data.get('keyName', '').strip()
        
        if not user_id or not key_name:
            return jsonify({'success': False, 'detail': 'User ID and key name required'}), 400
        
        # Verify user exists
        users_df = read_excel_data(USERS_EXCEL)
        user_row = users_df[users_df['id'] == user_id]
        if user_row.empty:
            return jsonify({'success': False, 'detail': 'User not found'}), 404
        
        # Generate new API key
        api_keys_df = read_excel_data(API_KEYS_EXCEL)
        api_key_id = generate_id()
        api_key = f"ak_{generate_id().replace('-', '')[:32]}"
        created_at = datetime.datetime.utcnow().isoformat()
        
        new_api_key = {
            'id': api_key_id,
            'user_id': user_id,
            'api_key': api_key,
            'name': key_name,
            'permissions': 'read,write',
            'active': 1,
            'created_at': created_at,
            'last_used': '',
            'usage_count': 0
        }
        
        api_keys_df = pd.concat([api_keys_df, pd.DataFrame([new_api_key])], ignore_index=True)
        if not write_excel_data(api_keys_df, API_KEYS_EXCEL):
            return jsonify({'success': False, 'detail': 'Failed to save API key'}), 500
        
        logger.info(f"API key generated: {key_name} for user {user_id}")
        
        return jsonify({
            'success': True,
            'message': f'API key "{key_name}" generated successfully!',
            'apiKey': {
                'id': api_key_id,
                'apiKey': api_key,
                'name': key_name,
                'createdAt': created_at,
                'active': True,
                'usageCount': 0
            }
        })
        
    except Exception as e:
        logger.error(f"API key generation error: {str(e)}")
        return jsonify({'success': False, 'detail': 'Failed to generate API key'}), 500

@app.route('/api/keys/<user_id>', methods=['GET'])
def get_user_api_keys(user_id):
    try:
        api_keys_df = read_excel_data(API_KEYS_EXCEL)
        user_keys = api_keys_df[api_keys_df['user_id'] == user_id]
        
        api_keys = []
        for _, key in user_keys.iterrows():
            api_keys.append({
                'id': key['id'],
                'apiKey': key['api_key'],
                'name': key['name'],
                'active': bool(key['active']),
                'createdAt': key['created_at'],
                'usageCount': key['usage_count'] if pd.notna(key['usage_count']) else 0
            })
        
        return jsonify({'success': True, 'apiKeys': api_keys})
        
    except Exception as e:
        logger.error(f"Get API keys error: {str(e)}")
        return jsonify({'success': False, 'detail': 'Failed to load API keys'}), 500

@app.route('/api/keys/<key_id>', methods=['DELETE'])
def revoke_api_key(key_id):
    try:
        data = request.get_json()
        user_id = data.get('userId')
        
        if not user_id:
            return jsonify({'success': False, 'detail': 'User ID required'}), 400
        
        api_keys_df = read_excel_data(API_KEYS_EXCEL)
        
        # Find and remove the key
        key_exists = ((api_keys_df['id'] == key_id) & (api_keys_df['user_id'] == user_id)).any()
        
        if not key_exists:
            return jsonify({'success': False, 'detail': 'API key not found or access denied'}), 404
        
        # Remove the key
        api_keys_df = api_keys_df[~((api_keys_df['id'] == key_id) & (api_keys_df['user_id'] == user_id))]
        
        if not write_excel_data(api_keys_df, API_KEYS_EXCEL):
            return jsonify({'success': False, 'detail': 'Failed to revoke API key'}), 500
        
        logger.info(f"API key revoked: {key_id} by user {user_id}")
        
        return jsonify({'success': True, 'message': 'API key revoked successfully'})
        
    except Exception as e:
        logger.error(f"API key revocation error: {str(e)}")
        return jsonify({'success': False, 'detail': 'Failed to revoke API key'}), 500

# Health check endpoint
@app.route('/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'healthy',
        'service': 'Aviality Healthcare Data Deidentification with Excel Storage',
        'timestamp': datetime.datetime.utcnow().isoformat(),
        'pii_detection_available': PII_DETECTION_AVAILABLE,
        'storage_type': 'Excel Files',
        'excel_files': {
            'users': os.path.exists(USERS_EXCEL),
            'api_keys': os.path.exists(API_KEYS_EXCEL),
            'processing_sessions': os.path.exists(PROCESSING_SESSIONS_EXCEL),
            'processed_data': os.path.exists(PROCESSED_DATA_EXCEL)
        }
    })

# Excel Data Export Endpoints for Admin/Debug
@app.route('/api/export-excel/<table_name>', methods=['GET'])
def export_excel_table(table_name):
    """Export Excel tables for viewing/backup"""
    try:
        file_mapping = {
            'users': USERS_EXCEL,
            'api_keys': API_KEYS_EXCEL,
            'processing_sessions': PROCESSING_SESSIONS_EXCEL,
            'processed_data': PROCESSED_DATA_EXCEL
        }
        
        if table_name not in file_mapping:
            return jsonify({'success': False, 'detail': 'Invalid table name'}), 400
        
        file_path = file_mapping[table_name]
        
        if not os.path.exists(file_path):
            return jsonify({'success': False, 'detail': 'Excel file not found'}), 404
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=f'{table_name}_export_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Excel export error: {str(e)}")
        return jsonify({'success': False, 'detail': 'Export failed'}), 500

# Error handlers
@app.errorhandler(413)
def too_large(e):
    return jsonify({'success': False, 'detail': 'File too large. Maximum size is 100MB.'}), 413

@app.errorhandler(500)
def internal_error(e):
    logger.error(f"Internal server error: {str(e)}")
    return jsonify({'success': False, 'detail': 'Internal server error occurred.'}), 500

if __name__ == '__main__':
    # Create necessary directories
    os.makedirs('uploads', exist_ok=True)
    os.makedirs(EXCEL_DATA_DIR, exist_ok=True)
    
    logger.info("🚀 Aviality Hospital System starting with Excel Storage...")
    logger.info("📊 Excel-based data storage for prototype phase")
    logger.info("🏥 User registration and login with Excel backend")
    logger.info("🔑 API key management stored in Excel")
    logger.info("📋 PII/PHI data structured as columns in Excel output")
    logger.info("💾 All data stored in excel_data/ directory")
    logger.info("🗄️ Easy data inspection and management")
    
    if PII_DETECTION_AVAILABLE:
        logger.info("🤖 Real PII/PHI detection enabled")
    else:
        logger.info("⚠️ Using sample hospital data")
    
    # Print Excel file locations
    logger.info(f"📁 Excel Storage Directory: {EXCEL_DATA_DIR}")
    logger.info(f"👥 Users: {USERS_EXCEL}")
    logger.info(f"🔑 API Keys: {API_KEYS_EXCEL}")
    logger.info(f"📊 Sessions: {PROCESSING_SESSIONS_EXCEL}")
    logger.info(f"💻 Processed Data: {PROCESSED_DATA_EXCEL}")
    
    app.run(debug=True, host='0.0.0.0', port=5000)

    # Add new Excel file for redactions log
REDACTIONS_LOG_EXCEL = os.path.join(EXCEL_DATA_DIR, 'redactions_log.xlsx')

def init_excel_storage():
    """Initialize Excel files for data storage"""
    os.makedirs(EXCEL_DATA_DIR, exist_ok=True)
    
    # ... existing code ...
    
    # Initialize Redactions Log Excel
    if not os.path.exists(REDACTIONS_LOG_EXCEL):
        redactions_df = pd.DataFrame(columns=[
            'id', 'session_id', 'user_id', 'filename', 'processed_at',
            'Hospital_ID', 'Patient_Name', 'Policy_Number', 'date_of_birth', 
            'email', 'health_insurance', 'phone_number',
            'Age', 'allergy', 'blood_group', 'drug', 
            'gender', 'medical_condition', 'surgery', 'symptom'
        ])
        redactions_df.to_excel(REDACTIONS_LOG_EXCEL, index=False)
        logger.info("Created redactions_log.xlsx")

@app.route('/api/redactions-history/<user_id>', methods=['GET'])
def get_redactions_history(user_id):
    try:
        redactions_df = read_excel_data(REDACTIONS_LOG_EXCEL)
        
        if redactions_df.empty:
            return jsonify({'success': True, 'history': []})
        
        # Filter by user if not admin
        user_redactions = redactions_df[redactions_df['user_id'] == user_id]
        
        history = []
        for _, record in user_redactions.iterrows():
            history.append({
                'id': record['id'],
                'filename': record['filename'],
                'processedAt': record['processed_at'],
                'sessionId': record['session_id']
            })
        
        return jsonify({'success': True, 'history': history})
        
    except Exception as e:
        logger.error(f"Get redactions history error: {str(e)}")
        return jsonify({'success': False, 'detail': 'Failed to load history'}), 500